from datetime import datetime
import re
import unicodedata

import pandas as pd
from openpyxl import load_workbook


ARCHIVO_ENTRADA = "ellagar_catalogo_completo.xlsx"
ARCHIVO_SALIDA = "proyecta_catalogo_normalizado.xlsx"


def convertir_a_slug(texto):
    if pd.isna(texto):
        return ""

    texto = str(texto).lower()

    texto = unicodedata.normalize("NFD", texto)

    texto = "".join(
        caracter
        for caracter in texto
        if unicodedata.category(caracter) != "Mn"
    )

    texto = re.sub(r"[^a-z0-9]+", "-", texto)
    texto = texto.strip("-")

    return texto


def construir_url_producto(producto_id, nombre):
    slug = convertir_a_slug(nombre)

    return (
        "https://www.ellagar.com/ECOMMERCE/"
        f"DetalleArticulo/{producto_id}/{slug}"
    )


def convertir_urls_en_links(archivo_excel):
    workbook = load_workbook(archivo_excel)
    sheet = workbook.active

    columna_url_producto = None
    columna_url_imagen = None

    for celda in sheet[1]:
        if celda.value == "URL_Producto":
            columna_url_producto = celda.column

        if celda.value == "URL_Imagen":
            columna_url_imagen = celda.column

    if columna_url_producto:
        for fila in range(2, sheet.max_row + 1):
            celda = sheet.cell(
                row=fila,
                column=columna_url_producto,
            )

            if celda.value:
                celda.hyperlink = celda.value
                celda.style = "Hyperlink"

    if columna_url_imagen:
        for fila in range(2, sheet.max_row + 1):
            celda = sheet.cell(
                row=fila,
                column=columna_url_imagen,
            )

            if celda.value:
                celda.hyperlink = celda.value
                celda.style = "Hyperlink"

    workbook.save(archivo_excel)


def main():
    print("Leyendo catálogo de El Lagar...")

    df = pd.read_excel(ARCHIVO_ENTRADA)

    fecha_actualizacion = datetime.now().strftime(
        "%Y-%m-%d %H:%M:%S"
    )

    catalogo = pd.DataFrame({
        "Proveedor": ["El Lagar"] * len(df),
        "ID_Proveedor": df["ID"],
        "SKU": df["Codigo_Principal"],
        "Nombre": df["Nombre"],
        "Marca": df["Marca"],
        "Categoria": df["CategoriaNombre"],
        "Subcategoria": df["SubCategoria"],
        "Precio_CRC": df["Precio"],
        "IVA": df["IVA"],
        "Es_Oferta": df["EsOferta"],
        "CABYS": df["CABYS"],
        "Descripcion": df["Descripcion"],
        "URL_Imagen": df["URLImagen"],
        "Compra_Online": df["Permite_Compra_Ecommerce"],
        "Fecha_Actualizacion": [fecha_actualizacion] * len(df),
    })

    catalogo["URL_Producto"] = catalogo.apply(
        lambda fila: construir_url_producto(
            fila["ID_Proveedor"],
            fila["Nombre"],
        ),
        axis=1,
    )

    catalogo.to_excel(
        ARCHIVO_SALIDA,
        index=False,
    )

    convertir_urls_en_links(ARCHIVO_SALIDA)

    print(f"Listo. Se creó: {ARCHIVO_SALIDA}")
    print(f"Productos normalizados: {len(catalogo)}")
    print("Las URLs de producto e imagen ahora son clickeables.")


if __name__ == "__main__":
    main()